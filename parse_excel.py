import openpyxl
import json

wb = openpyxl.load_workbook(r"c:\Projects\STS_quiz\STS 4006 MCQ's.xlsx")

data = {}
for sn in wb.sheetnames:
    ws = wb[sn]
    questions = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        vals = [c.value for c in row[:8]]
        if vals[2]:  # has question
            questions.append({
                "topic": str(vals[0]).strip() if vals[0] else "General",
                "difficulty": str(vals[1]).strip() if vals[1] else "Medium",
                "question": str(vals[2]).strip(),
                "options": [
                    str(vals[3]).strip() if vals[3] else "",
                    str(vals[4]).strip() if vals[4] else "",
                    str(vals[5]).strip() if vals[5] else "",
                    str(vals[6]).strip() if vals[6] else "",
                ],
                "answer": str(vals[7]).strip().upper() if vals[7] else "A"
            })
    data[sn] = questions

js = "const QUIZ_DATA = " + json.dumps(data, indent=2, ensure_ascii=False) + ";\n"

with open(r"c:\Projects\STS_quiz\data.js", "w", encoding="utf-8") as f:
    f.write(js)

for k, v in data.items():
    print(f"{k}: {len(v)} questions")
print("data.js written successfully")
